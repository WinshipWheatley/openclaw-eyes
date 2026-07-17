"""Trusted workbook reconciliation, verification, and invoice package publication.

This module owns no draft, send, payment, or ledger authority. It edits only an
isolated workbook copy and publishes a package after deterministic verification.
"""

from __future__ import annotations

import calendar
import hashlib
import json
import os
import re
import shutil
import subprocess
import uuid
from copy import copy
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string


SCHEMA_VERSION = "invoice_workbook_finalizer_v1"
MANIFEST_SCHEMA = "openclaw_invoice_manifest_v1"
POWERSHELL_EXE = Path("/mnt/c/Windows/System32/WindowsPowerShell/v1.0/powershell.exe")
RECALCULATE_SCRIPT = Path(__file__).resolve().parent / "scripts" / "recalculate_invoice_with_excel.ps1"
EXPORT_PDF_SCRIPT = Path(__file__).resolve().parent / "scripts" / "export_invoice_pdf_with_excel.ps1"
SEMANTIC_MARKERS = (
    ("DRAFT", re.compile(r"\bdraft\b", re.IGNORECASE)),
    ("TBD", re.compile(r"\btbd\b", re.IGNORECASE)),
    ("TODO", re.compile(r"\bto\s*do\b", re.IGNORECASE)),
    ("PLACEHOLDER", re.compile(r"\bplaceholder\b|\badd confirmed\b", re.IGNORECASE)),
)
ABSOLUTE_PRINT_AREA = re.compile(
    r"^\$([A-Z]{1,3})\$([1-9]\d*):\$([A-Z]{1,3})\$([1-9]\d*)$"
)


class InvoiceFinalizationError(RuntimeError):
    """Fail-closed W1 verification error with a stable machine code."""


def _validated_print_area(value: str) -> str:
    match = ABSOLUTE_PRINT_AREA.fullmatch(value)
    if match is None:
        raise InvoiceFinalizationError("PRINT_AREA_INVALID")
    start_column, start_row, end_column, end_row = match.groups()
    start_column_index = column_index_from_string(start_column)
    end_column_index = column_index_from_string(end_column)
    if (
        start_column_index > end_column_index
        or int(start_row) > int(end_row)
        or end_column_index > 16384
        or int(end_row) > 1048576
    ):
        raise InvoiceFinalizationError("PRINT_AREA_INVALID")
    return value


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _stable_json(payload: object) -> str:
    return json.dumps(payload, sort_keys=True, separators=(",", ":"), default=str)


def _formula_map(path: str | Path) -> dict[str, str]:
    book = load_workbook(Path(path), data_only=False, read_only=True)
    formulas: dict[str, str] = {}
    for sheet in book.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas[f"{sheet.title}!{cell.coordinate}"] = cell.value
    return formulas


def formula_sha256(path: str | Path) -> str:
    return hashlib.sha256(_stable_json(_formula_map(path)).encode("utf-8")).hexdigest()


def semantic_marker_findings(path: str | Path, sheet_name: str) -> list[dict[str, str]]:
    book = load_workbook(Path(path), data_only=False, read_only=True)
    if sheet_name not in book.sheetnames:
        raise InvoiceFinalizationError(f"SOURCE_SHEET_MISSING:{sheet_name}")
    findings: list[dict[str, str]] = []
    for row in book[sheet_name].iter_rows():
        for cell in row:
            if not isinstance(cell.value, str) or cell.value.startswith("="):
                continue
            for marker, pattern in SEMANTIC_MARKERS:
                if pattern.search(cell.value):
                    findings.append({"cell": cell.coordinate, "marker": marker})
                    break
    return findings


def _clear_invoice_lines(sheet: Any) -> None:
    for row in range(22, 34):
        for column in range(2, 7):
            sheet.cell(row=row, column=column).value = None


def _set_register_row(
    sheet: Any,
    row: int,
    source_sheet: str,
    work_type: str,
    note: str,
    *,
    status: str | None = None,
) -> None:
    values = (
        f"='{source_sheet}'!$G$3",
        source_sheet,
        "Live Arts Maryland",
        work_type,
        f"='{source_sheet}'!$G$4",
        f"='{source_sheet}'!$G$43",
        f"='{source_sheet}'!$E$50",
        f"='{source_sheet}'!$G$50",
        f"='{source_sheet}'!$C$51",
        status if status is not None else f"='{source_sheet}'!$C$50",
        note,
    )
    for column, value in enumerate(values, start=1):
        sheet.cell(row=row, column=column, value=value)


def apply_lamd_truth_reconciliation(
    source_path: str | Path,
    output_path: str | Path,
) -> dict[str, Any]:
    """Apply operator-confirmed LAMD facts to an isolated workbook copy."""

    source = Path(source_path)
    output = Path(output_path)
    if source.resolve() == output.resolve():
        raise InvoiceFinalizationError("SOURCE_COPY_REQUIRED")
    if not source.is_file() or source.stat().st_size == 0:
        raise InvoiceFinalizationError("SOURCE_WORKBOOK_MISSING")
    source_hash = _sha256(source)
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, output)
    book = load_workbook(output, data_only=False, read_only=False)
    required = {"June 2026 Speaker Rental", "June 2026 AV Tech", "July 2026", "Invoice Register"}
    missing = sorted(required - set(book.sheetnames))
    if missing:
        raise InvoiceFinalizationError(f"SOURCE_SHEETS_MISSING:{','.join(missing)}")
    if "June 2026 Rental" in book.sheetnames:
        raise InvoiceFinalizationError("JUNE_RECONCILIATION_ALREADY_PRESENT")

    paid_1001 = book["June 2026 Speaker Rental"]
    paid_1001["C50"] = "PAID - direct deposit received 2026-06-09"
    paid_1001["E50"] = 900
    paid_1001["E51"] = date(2026, 6, 9)
    paid_1001["G51"] = "Operator-confirmed payment; confirmation reference retained outside workbook receipts."

    july = book["July 2026"]
    june = book.copy_worksheet(paid_1001)
    june.title = "June 2026 Rental"
    _clear_invoice_lines(june)
    june["G2"] = "INVOICE"
    june["G3"] = "2026-1003"
    june["G4"] = date(2026, 6, 23)
    june["C5"] = "Speaker rental - June 2026"
    june["G5"] = "Live Arts Maryland / Speaker Rentals"
    june["B22"] = 1
    june["C22"] = "Speaker rental - June 2026"
    june["D22"] = 1
    june["E22"] = 100
    june["F22"] = 0
    june["C50"] = "SENT 2026-06-23"
    june["E50"] = 100
    june["E51"] = date(2026, 6, 25)
    june["G51"] = "Payment confirmed 2026-06-24; direct deposit received 2026-06-25."

    _clear_invoice_lines(july)
    july["G2"] = "INVOICE"
    july["G3"] = "2026-1004"
    july["G4"] = date(2026, 7, 16)
    july["C5"] = "Speaker rental - July 2026"
    july["G5"] = "Live Arts Maryland / Speaker Rentals"
    july["B22"] = 1
    july["C22"] = "Speaker rental - July 2026"
    july["D22"] = 1
    july["E22"] = 100
    july["F22"] = 0
    july["C48"] = "Thank you - July 2026 speaker rental invoice."
    july["B52"] = "='June 2026 Rental'!$G$3"
    july["C52"] = "Speaker rental"
    july["D52"] = "='June 2026 Rental'!$G$43"
    july["E52"] = "='June 2026 Rental'!$E$50"
    july["F52"] = "='June 2026 Rental'!$G$50"
    july["G52"] = "='June 2026 Rental'!$C$51"

    register = book["Invoice Register"]
    register.insert_rows(7, amount=1)
    for column in range(1, 12):
        register.cell(row=7, column=column)._style = copy(register.cell(row=8, column=column)._style)
        register.cell(row=7, column=column).number_format = register.cell(row=8, column=column).number_format
    register["J6"] = "ARCHIVED - UNISSUED CANDIDATE"
    register["K6"] = "Reserved invoice number retained; candidate is not a live receivable."
    _set_register_row(
        register,
        7,
        "June 2026 Rental",
        "Speaker rental",
        "Sent 2026-06-23; paid 2026-06-25; operator-confirmed correction.",
    )
    _set_register_row(
        register,
        8,
        "July 2026",
        "Speaker rental",
        "Finalized verification target; SEND_HOLD; no draft or send performed.",
        status="FINALIZED - SEND_HOLD",
    )
    book.save(output)
    if _sha256(source) != source_hash:
        raise InvoiceFinalizationError("SOURCE_MUTATED")
    return {
        "schema_version": SCHEMA_VERSION,
        "status": "RECONCILED_COPY_READY_FOR_EXCEL",
        "source_sha256": source_hash,
        "reconciled_sha256": _sha256(output),
        "pre_excel_formula_sha256": formula_sha256(output),
        "source_mutated": False,
        "numbers": {
            "paid_sep_2025_through_may_2026": "2026-1001",
            "existing_av_tech_reservation": "2026-1002",
            "paid_june_2026": "2026-1003",
            "live_july_2026": "2026-1004",
        },
        "authority_boundary": {
            "draft_created": False,
            "external_send_performed": False,
            "money_moved": False,
            "ledger_posted": False,
        },
    }


def _decimal_number(value: object, *, code: str) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise InvoiceFinalizationError(code)
    return float(value)


def verify_recalculated_invoice(
    workbook_path: str | Path,
    *,
    sheet_name: str,
    invoice_number: str,
    expected_amount: float,
    expected_formula_sha256: str,
    balance_cell: str = "G50",
) -> dict[str, Any]:
    path = Path(workbook_path)
    findings = semantic_marker_findings(path, sheet_name)
    if findings:
        raise InvoiceFinalizationError("SEMANTIC_MARKER_FOUND")
    actual_formula_hash = formula_sha256(path)
    if actual_formula_hash != expected_formula_sha256:
        raise InvoiceFinalizationError("FORMULA_MUTATION")

    formulas = load_workbook(path, data_only=False, read_only=True)
    values = load_workbook(path, data_only=True, read_only=True)
    formula_sheet = formulas[sheet_name]
    value_sheet = values[sheet_name]
    if str(formula_sheet["G3"].value or "").strip() != invoice_number:
        raise InvoiceFinalizationError("INVOICE_NUMBER_MISMATCH")
    critical_cells = tuple(dict.fromkeys(("G22", "G40", "G41", "G42", "G43", balance_cell)))
    critical_values = {cell: value_sheet[cell].value for cell in critical_cells}
    if any(value is None for value in critical_values.values()):
        raise InvoiceFinalizationError("STALE_FORMULA_CACHE")

    independent_subtotal = 0.0
    for row in range(22, 34):
        quantity = formula_sheet.cell(row=row, column=4).value
        unit_price = formula_sheet.cell(row=row, column=5).value
        discount = formula_sheet.cell(row=row, column=6).value
        if quantity in (None, "") and unit_price in (None, ""):
            continue
        independent_subtotal += _decimal_number(quantity or 0, code="QUANTITY_INVALID") * _decimal_number(
            unit_price or 0, code="UNIT_PRICE_INVALID"
        ) - _decimal_number(discount or 0, code="DISCOUNT_INVALID")
    cached_subtotal = _decimal_number(critical_values["G40"], code="CACHED_SUBTOTAL_INVALID")
    cached_total = _decimal_number(critical_values["G43"], code="CACHED_TOTAL_INVALID")
    cached_balance = _decimal_number(
        critical_values[balance_cell], code="CACHED_BALANCE_INVALID"
    )
    expected = float(expected_amount)
    if any(abs(value - expected) > 0.005 for value in (independent_subtotal, cached_subtotal, cached_total, cached_balance)):
        raise InvoiceFinalizationError("INDEPENDENT_TOTAL_MISMATCH")
    return {
        "schema_version": SCHEMA_VERSION,
        "receipt_id": f"invoice-w1-verify:{_sha256(path)[:24]}",
        "formula_freshness_receipt_id": f"invoice-w1-formula:{actual_formula_hash[:24]}",
        "status": "VERIFIED",
        "workbook_sha256": _sha256(path),
        "formula_sha256": actual_formula_hash,
        "sheet_name": sheet_name,
        "invoice_number": invoice_number,
        "amount": expected,
        "critical_cache_cells_verified": list(critical_cells),
        "semantic_marker_findings": [],
        "independent_recompute": {
            "subtotal": independent_subtotal,
            "total": cached_total,
            "balance_due": cached_balance,
            "matches": True,
        },
        "authority_boundary": {
            "draft_created": False,
            "external_send_performed": False,
            "money_moved": False,
            "ledger_posted": False,
        },
    }


def _windows_path(path: Path) -> str:
    completed = subprocess.run(
        ["wslpath", "-w", str(path.resolve())],
        check=False,
        capture_output=True,
        text=True,
        timeout=10,
    )
    if completed.returncode != 0 or not completed.stdout.strip():
        raise InvoiceFinalizationError("WINDOWS_PATH_CONVERSION_FAILED")
    return completed.stdout.strip()


def _run_excel_worker(script: Path, arguments: list[str], *, timeout: int) -> dict[str, Any]:
    if not POWERSHELL_EXE.is_file():
        raise InvoiceFinalizationError("POWERSHELL_NOT_FOUND")
    if not script.is_file():
        raise InvoiceFinalizationError("EXCEL_WORKER_NOT_FOUND")
    command = [
        str(POWERSHELL_EXE),
        "-NoProfile",
        "-NonInteractive",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        _windows_path(script),
        *arguments,
    ]
    try:
        completed = subprocess.run(
            command,
            check=False,
            capture_output=True,
            text=True,
            timeout=timeout,
        )
    except subprocess.TimeoutExpired as exc:
        raise InvoiceFinalizationError("EXCEL_WORKER_TIMEOUT") from exc
    json_lines = [line.strip() for line in completed.stdout.splitlines() if line.strip().startswith("{")]
    if not json_lines:
        raise InvoiceFinalizationError("EXCEL_WORKER_RECEIPT_MISSING")
    try:
        receipt = json.loads(json_lines[-1])
    except json.JSONDecodeError as exc:
        raise InvoiceFinalizationError("EXCEL_WORKER_RECEIPT_INVALID") from exc
    if completed.returncode != 0 or receipt.get("status") == "FAILED":
        code = str(receipt.get("error_code") or "EXCEL_WORKER_FAILED")
        raise InvoiceFinalizationError(code)
    return receipt


def recalculate_with_excel(
    source_path: str | Path,
    output_path: str | Path,
    *,
    timeout: int = 180,
) -> dict[str, Any]:
    source = Path(source_path)
    output = Path(output_path)
    if source.resolve() == output.resolve():
        raise InvoiceFinalizationError("RECALC_OUTPUT_MUST_DIFFER")
    if not source.is_file() or source.stat().st_size == 0:
        raise InvoiceFinalizationError("SOURCE_WORKBOOK_MISSING")
    if output.exists():
        raise InvoiceFinalizationError("RECALC_OUTPUT_ALREADY_EXISTS")
    output.parent.mkdir(parents=True, exist_ok=True)
    before_formula_hash = formula_sha256(source)
    receipt = _run_excel_worker(
        RECALCULATE_SCRIPT,
        [
            "-InputPath",
            _windows_path(source),
            "-OutputPath",
            _windows_path(output),
            "-TimeoutSeconds",
            str(max(10, min(timeout - 10, 600))),
        ],
        timeout=timeout,
    )
    if not output.is_file() or output.stat().st_size == 0:
        raise InvoiceFinalizationError("RECALC_OUTPUT_MISSING")
    after_formula_hash = formula_sha256(output)
    if before_formula_hash != after_formula_hash:
        raise InvoiceFinalizationError("FORMULA_MUTATION")
    return {
        "schema_version": SCHEMA_VERSION,
        "status": "RECALCULATED",
        "source_sha256": _sha256(source),
        "output_sha256": _sha256(output),
        "formula_sha256": after_formula_hash,
        "excel_version": str(receipt.get("excel_version") or ""),
        "excel_build": str(receipt.get("excel_build") or ""),
        "calculation_state": str(receipt.get("calculation_state") or ""),
        "reopen_count": int(receipt.get("reopen_count") or 0),
        "macros_disabled": True,
        "external_links_allowed": False,
        "ambient_excel_processes_terminated": False,
    }


def export_invoice_pdf_with_excel(
    workbook_path: str | Path,
    *,
    sheet_name: str,
    output_path: str | Path,
    print_area: str | None = None,
    timeout: int = 180,
) -> dict[str, Any]:
    workbook = Path(workbook_path)
    output = Path(output_path)
    if not workbook.is_file() or workbook.stat().st_size == 0:
        raise InvoiceFinalizationError("WORKBOOK_MISSING")
    if output.exists():
        raise InvoiceFinalizationError("PDF_OUTPUT_ALREADY_EXISTS")
    bounded_print_area = _validated_print_area(print_area) if print_area else None
    workbook_sha256_before = _sha256(workbook)
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.parent / f".{output.name}.tmp.{uuid.uuid4().hex}.pdf"
    worker_args = [
        "-InputPath",
        _windows_path(workbook),
        "-SheetName",
        sheet_name,
        "-OutputPath",
        _windows_path(temporary),
    ]
    if bounded_print_area:
        worker_args.extend(["-PrintArea", bounded_print_area])
    receipt = _run_excel_worker(
        EXPORT_PDF_SCRIPT,
        worker_args,
        timeout=timeout,
    )
    workbook_sha256_after = _sha256(workbook)
    if workbook_sha256_after != workbook_sha256_before:
        temporary.unlink(missing_ok=True)
        raise InvoiceFinalizationError("WORKBOOK_MUTATED_DURING_EXPORT")
    excel_print_area = str(receipt.get("print_area") or "")
    fit_to_pages_wide = int(receipt.get("fit_to_pages_wide") or 0)
    fit_to_pages_tall = int(receipt.get("fit_to_pages_tall") or 0)
    if bounded_print_area and (
        not excel_print_area or fit_to_pages_wide != 1 or fit_to_pages_tall != 1
    ):
        temporary.unlink(missing_ok=True)
        raise InvoiceFinalizationError("PRINT_FRAME_NOT_APPLIED")
    proof = _validate_pdf(temporary)
    os.replace(temporary, output)
    return {
        "schema_version": SCHEMA_VERSION,
        "status": "PDF_EXPORTED_VERIFIED",
        "pdf_sha256": proof["sha256"],
        "pdf_size": proof["size"],
        "pdf_page_count": proof["page_count"],
        "excel_version": str(receipt.get("excel_version") or ""),
        "sheet_name": sheet_name,
        "print_area_applied": bounded_print_area,
        "excel_print_area_reported": excel_print_area,
        "fit_to_pages_wide": fit_to_pages_wide,
        "fit_to_pages_tall": fit_to_pages_tall,
        "workbook_sha256_before": workbook_sha256_before,
        "workbook_sha256_after": workbook_sha256_after,
        "workbook_unchanged": True,
        "atomic_publish": True,
        "external_send_performed": False,
    }


def _validate_pdf(path: Path) -> dict[str, Any]:
    if not path.is_file():
        raise InvoiceFinalizationError("PDF_MISSING")
    if path.stat().st_size == 0:
        raise InvoiceFinalizationError("PDF_ZERO_BYTE")
    if not path.read_bytes()[:5] == b"%PDF-":
        raise InvoiceFinalizationError("PDF_SIGNATURE_INVALID")
    pdfinfo = shutil.which("pdfinfo")
    if pdfinfo:
        completed = subprocess.run(
            [pdfinfo, str(path)],
            check=False,
            capture_output=True,
            text=True,
            timeout=20,
        )
        if completed.returncode != 0:
            raise InvoiceFinalizationError("PDF_STRUCTURE_INVALID")
        match = re.search(r"^Pages:\s+(\d+)$", completed.stdout, re.MULTILINE)
        page_count = int(match.group(1)) if match else 0
    else:
        page_count = 0
    pdftotext = shutil.which("pdftotext")
    if not pdftotext:
        raise InvoiceFinalizationError("PDF_TEXT_EXTRACTOR_MISSING")
    extracted = subprocess.run(
        [pdftotext, str(path), "-"],
        check=False,
        capture_output=True,
        text=True,
        timeout=20,
    )
    if extracted.returncode != 0:
        raise InvoiceFinalizationError("PDF_TEXT_EXTRACTION_FAILED")
    if any(pattern.search(extracted.stdout) for _, pattern in SEMANTIC_MARKERS):
        raise InvoiceFinalizationError("PDF_SEMANTIC_MARKER_FOUND")
    return {
        "sha256": _sha256(path),
        "size": path.stat().st_size,
        "page_count": page_count,
        "text_sha256": hashlib.sha256(extracted.stdout.encode("utf-8")).hexdigest(),
    }


def publish_verified_package(
    package_dir: str | Path,
    *,
    workbook_path: str | Path,
    pdf_path: str | Path,
    specification: Mapping[str, Any],
    verification_receipt: Mapping[str, Any],
) -> dict[str, Any]:
    package = Path(package_dir)
    workbook = Path(workbook_path)
    pdf = Path(pdf_path)
    if verification_receipt.get("status") != "VERIFIED":
        raise InvoiceFinalizationError("VERIFICATION_RECEIPT_REQUIRED")
    if not workbook.is_file() or workbook.stat().st_size == 0:
        raise InvoiceFinalizationError("WORKBOOK_MISSING")
    pdf_proof = _validate_pdf(pdf)
    required = ("client_ref", "invoice_number", "service_period", "source_sheet", "amount")
    missing = [key for key in required if specification.get(key) in (None, "")]
    if missing:
        raise InvoiceFinalizationError(f"SPECIFICATION_INCOMPLETE:{','.join(missing)}")
    service_period = str(specification["service_period"])
    if re.fullmatch(r"\d{4}-\d{2}", service_period) is None:
        raise InvoiceFinalizationError("SERVICE_PERIOD_INVALID")
    year, month = (int(part) for part in service_period.split("-"))
    if package.exists():
        raise InvoiceFinalizationError("PACKAGE_ALREADY_EXISTS")

    stage = package.parent / f".{package.name}.tmp.{uuid.uuid4().hex}"
    stage.mkdir(parents=True, exist_ok=False)
    target_workbook = stage / "invoice.xlsx"
    target_pdf = stage / "invoice.pdf"
    shutil.copy2(workbook, target_workbook)
    shutil.copy2(pdf, target_pdf)
    manifest = {
        "schema": MANIFEST_SCHEMA,
        "invoice_key": f"{service_period}_{specification['client_ref']}_{specification['invoice_number']}",
        "client_slug": str(specification["client_ref"]).replace("_", "-"),
        "invoice_number": str(specification["invoice_number"]),
        "service_period_start": f"{service_period}-01",
        "service_period_end": f"{service_period}-{calendar.monthrange(year, month)[1]:02d}",
        "status": "finalized_verified",
        "amount": float(specification["amount"]),
        "source_sheet": str(specification["source_sheet"]),
        "package_workbook_sha256": _sha256(target_workbook),
        "current_pdf_sha256": _sha256(target_pdf),
        "latest_send_receipt_path": None,
        "artifact_verification_receipt_id": str(verification_receipt.get("receipt_id") or ""),
        "formula_freshness_receipt_id": str(
            verification_receipt.get("formula_freshness_receipt_id") or ""
        ),
        "authority_boundary": {
            "provider_draft_created": False,
            "external_send_performed": False,
            "money_moved": False,
            "ledger_posted": False,
        },
    }
    (stage / "invoice_manifest.json").write_text(
        json.dumps(manifest, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    os.replace(stage, package)
    return {
        "schema_version": SCHEMA_VERSION,
        "status": "PUBLISHED_VERIFIED",
        "published_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "package_path": package.as_posix(),
        "workbook_sha256": manifest["package_workbook_sha256"],
        "pdf_sha256": manifest["current_pdf_sha256"],
        "pdf_size": pdf_proof["size"],
        "pdf_page_count": pdf_proof["page_count"],
        "manifest": manifest,
        "authority_boundary": manifest["authority_boundary"],
    }
