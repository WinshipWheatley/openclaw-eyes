from __future__ import annotations

import hashlib
import json
from pathlib import Path

import openpyxl

import st_annes_invoice_truth_drift as drift


FIXED_NOW = "2026-07-16T17:00:00+00:00"


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _fixture(tmp_path: Path, *, sheet_name: str = "June 2026") -> tuple[Path, Path, Path]:
    package_dir = tmp_path / "st-annes-2026-06"
    package_dir.mkdir()
    workbook_path = package_dir / "invoice.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet["G3"] = "3"
    sheet["J3"] = "Draft"
    sheet["G5"] = "A/V Tech - June 2026 Services"
    sheet["B21"] = "Item #"
    sheet["C21"] = "Description"
    sheet["D21"] = "Qty"
    sheet["E21"] = "Unit price"
    sheet["G21"] = "Running Total"
    running = 0.0
    for offset, day in enumerate((7, 10, 13, 14, 21, 27, 28), start=1):
        row = 21 + offset
        running += 125.0
        sheet.cell(row, 2, offset)
        sheet.cell(row, 3, f"June {day}, 2026 service")
        sheet.cell(row, 4, 1)
        sheet.cell(row, 5, 125.0)
        sheet.cell(row, 7, running)
    sheet["F36"] = "TOTAL DUE"
    sheet["G36"] = 875.0
    workbook.save(workbook_path)

    manifest_path = package_dir / "invoice_manifest.json"
    manifest_path.write_text(
        json.dumps(
            {
                "schema": "openclaw_invoice_manifest_v1",
                "invoice_key": "2026-06_st-annes",
                "client_slug": "st-annes",
                "invoice_number": "3",
                "service_period_start": "2026-06-01",
                "service_period_end": "2026-06-30",
                "status": "draft",
                "amount": 875.0,
                "source_sheet": "June 2026",
                "package_workbook_sha256": _sha256(workbook_path),
                "latest_send_receipt_path": None,
            }
        )
        + "\n",
        encoding="utf-8",
    )
    hygiene_path = tmp_path / "st_annes_work_log_hygiene.json"
    hygiene_path.write_text(
        json.dumps({"business_confirmed_ready_event_ids": []}) + "\n",
        encoding="utf-8",
    )
    return manifest_path, hygiene_path, workbook_path


def test_detects_seven_workbook_services_against_zero_confirmed_mirror(tmp_path: Path) -> None:
    manifest_path, hygiene_path, workbook_path = _fixture(tmp_path)
    before = _sha256(workbook_path)

    payload = drift.build_truth_drift(
        manifest_path,
        hygiene_path,
        generated_at=FIXED_NOW,
    )

    assert payload["status"] == "DRIFT_DETECTED"
    assert payload["workbook_truth"]["invoice_number"] == "3"
    assert payload["workbook_truth"]["source_sheet"] == "June 2026"
    assert payload["workbook_truth"]["service_count"] == 7
    assert payload["workbook_truth"]["total_due"] == 875.0
    assert payload["mirror_truth"]["confirmed_event_count"] == 0
    assert payload["missing_items"] == [
        "Reconcile workbook billables into the work-log mirror"
    ]
    assert payload["machine_proof"]["workbook_hash_verified"] is True
    assert payload["machine_proof"]["workbook_mutation_performed"] is False
    assert payload["machine_proof"]["ledger_mutation_performed"] is False
    assert payload["machine_proof"]["email_send_performed"] is False
    assert _sha256(workbook_path) == before


def test_hash_mismatch_fails_closed_before_workbook_parse(tmp_path: Path) -> None:
    manifest_path, hygiene_path, _workbook_path = _fixture(tmp_path)
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["package_workbook_sha256"] = "0" * 64
    manifest_path.write_text(json.dumps(manifest) + "\n", encoding="utf-8")

    payload = drift.build_truth_drift(manifest_path, hygiene_path, generated_at=FIXED_NOW)

    assert payload["status"] == "HASH_MISMATCH"
    assert payload["workbook_truth"] == {}
    assert payload["machine_proof"]["workbook_hash_verified"] is False


def test_missing_declared_sheet_and_malformed_mirror_fail_closed(tmp_path: Path) -> None:
    manifest_path, hygiene_path, _workbook_path = _fixture(tmp_path, sheet_name="May 2026")
    hygiene_path.write_text(
        json.dumps({"business_confirmed_ready_event_ids": "not-a-list"}) + "\n",
        encoding="utf-8",
    )

    payload = drift.build_truth_drift(manifest_path, hygiene_path, generated_at=FIXED_NOW)

    assert payload["status"] == "SOURCE_UNAVAILABLE"
    assert "declared_sheet_missing:June 2026" in payload["blockers"]
    assert "mirror_confirmed_ids_invalid" in payload["blockers"]
    assert payload["machine_proof"]["workbook_mutation_performed"] is False
