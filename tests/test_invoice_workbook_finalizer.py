from __future__ import annotations

import hashlib
import json
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from reportlab.pdfgen.canvas import Canvas

import invoice_artifact_locator
import invoice_workbook_finalizer as finalizer


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _invoice_sheet(book: Workbook, title: str, number: str | None, amount: int) -> None:
    sheet = book.create_sheet(title)
    sheet["G2"] = "INVOICE"
    sheet["F3"] = "Invoice #:"
    sheet["G3"] = number or "=\"2026-\"&TEXT(1003,\"0000\")"
    sheet["F4"] = "Date:"
    sheet["G4"] = "=TODAY()"
    sheet["F5"] = "Job:"
    sheet["G5"] = "Live Arts Maryland / Speaker Rentals"
    sheet["C22"] = "Speaker rental"
    sheet["D22"] = 1 if amount else 0
    sheet["E22"] = amount
    sheet["F22"] = 0
    sheet["G22"] = "=D22*E22-F22"
    sheet["G40"] = "=MAX(G22:G33)"
    sheet["G41"] = "=0"
    sheet["G42"] = "=G40*G41"
    sheet["G43"] = "=G40+G42"
    sheet["B50"] = "Invoice Status"
    sheet["C50"] = "Draft - ready to send"
    sheet["D50"] = "Amount Received"
    sheet["E50"] = 0
    sheet["F50"] = "Balance Due"
    sheet["G50"] = "=MAX(0,G43-E50)"
    sheet["B51"] = "Receipt Status"
    sheet["C51"] = '=IF(G50=0,"PAID",IF(E50>0,"PARTIAL","UNPAID"))'
    sheet["D51"] = "Payment Date"
    sheet["E51"] = None
    sheet["F51"] = "Ledger Match"
    sheet["G51"] = "Pending"


def _lamd_fixture(path: Path) -> None:
    book = Workbook()
    book.remove(book.active)
    _invoice_sheet(book, "June 2026 Speaker Rental", "2026-1001", 900)
    _invoice_sheet(book, "June 2026 AV Tech", "2026-1002", 4625)
    _invoice_sheet(book, "July 2026", None, 0)
    july = book["July 2026"]
    july["G2"] = "JULY DRAFT"
    july["C22"] = "July service line - add confirmed description"
    july["C48"] = "future invoice draft; check prior receipts before carrying a balance"
    july["B49"] = "PRIOR INVOICE RECEIPT CHECK"
    for cell, value in {
        "B50": "Source Invoice",
        "C50": "Work Type",
        "D50": "Invoice Amount",
        "E50": "Amount Received",
        "F50": "Balance Due",
        "G50": "Receipt Status",
        "B51": "='June 2026 Speaker Rental'!$G$3",
        "C51": "Speaker rental",
        "D51": "='June 2026 Speaker Rental'!$G$43",
        "E51": "='June 2026 Speaker Rental'!$E$50",
        "F51": "='June 2026 Speaker Rental'!$G$50",
        "G51": "='June 2026 Speaker Rental'!$C$51",
        "B52": "='June 2026 AV Tech'!$G$3",
        "C52": "A/V tech work",
        "D52": "='June 2026 AV Tech'!$G$43",
        "E52": "='June 2026 AV Tech'!$E$50",
        "F52": "='June 2026 AV Tech'!$G$50",
        "G52": "='June 2026 AV Tech'!$C$51",
    }.items():
        july[cell] = value

    register = book.create_sheet("Invoice Register")
    headers = [
        "Invoice #",
        "Sheet",
        "Client",
        "Work Type",
        "Invoice Date",
        "Invoice Amount",
        "Amount Received",
        "Balance Due",
        "Receipt Status",
        "Invoice Status",
        "Notes / Ledger Match",
    ]
    for index, value in enumerate(headers, start=1):
        register.cell(row=4, column=index, value=value)
    for row, sheet_name in ((5, "June 2026 Speaker Rental"), (6, "June 2026 AV Tech"), (7, "July 2026")):
        register.cell(row=row, column=1, value=f"='{sheet_name}'!$G$3")
        register.cell(row=row, column=2, value=sheet_name)
        register.cell(row=row, column=3, value="Live Arts Maryland")
        register.cell(row=row, column=4, value="Speaker rental")
        register.cell(row=row, column=5, value=f"='{sheet_name}'!$G$4")
        register.cell(row=row, column=6, value=f"='{sheet_name}'!$G$43")
        register.cell(row=row, column=7, value=f"='{sheet_name}'!$E$50")
        register.cell(row=row, column=8, value=f"='{sheet_name}'!$G$50")
        register.cell(row=row, column=9, value=f"='{sheet_name}'!$C$51")
        register.cell(row=row, column=10, value=f"='{sheet_name}'!$C$50")
    book.save(path)


def test_lamd_truth_batch_repairs_paid_rows_and_reserves_unique_july_number(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    reconciled = tmp_path / "reconciled.xlsx"
    _lamd_fixture(source)
    source_hash = _sha256(source)

    receipt = finalizer.apply_lamd_truth_reconciliation(source, reconciled)

    assert _sha256(source) == source_hash
    book = load_workbook(reconciled, data_only=False)
    paid_1001 = book["June 2026 Speaker Rental"]
    assert paid_1001["E50"].value == 900
    assert paid_1001["E51"].value.date() == date(2026, 6, 9)
    assert "Draft" not in str(paid_1001["C50"].value)

    june = book["June 2026 Rental"]
    assert june["G3"].value == "2026-1003"
    assert june["E22"].value == 100
    assert june["E50"].value == 100
    assert june["E51"].value.date() == date(2026, 6, 25)
    assert "2026-06-23" in str(june["C50"].value)

    july = book["July 2026"]
    assert july["G3"].value == "2026-1004"
    assert july["G4"].value.date() == date(2026, 7, 16)
    assert july["D22"].value == 1
    assert july["E22"].value == 100
    assert july["C48"].value == "Thank you - July 2026 speaker rental invoice."
    assert july["B50"].value == "Source Invoice"
    assert july["G50"].value == "Receipt Status"
    assert july["B51"].value == "='June 2026 Speaker Rental'!$G$3"
    assert july["B52"].value == "='June 2026 Rental'!$G$3"
    assert finalizer.semantic_marker_findings(reconciled, "July 2026") == []

    register = book["Invoice Register"]
    assert register["J6"].value == "ARCHIVED - UNISSUED CANDIDATE"
    assert register["J8"].value == "FINALIZED - SEND_HOLD"

    assert receipt["status"] == "RECONCILED_COPY_READY_FOR_EXCEL"
    assert receipt["source_mutated"] is False
    assert receipt["numbers"] == {
        "paid_sep_2025_through_may_2026": "2026-1001",
        "existing_av_tech_reservation": "2026-1002",
        "paid_june_2026": "2026-1003",
        "live_july_2026": "2026-1004",
    }


def test_semantic_markers_fail_before_finalization(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _lamd_fixture(source)

    findings = finalizer.semantic_marker_findings(source, "July 2026")

    assert {item["cell"] for item in findings} >= {"G2", "C22", "C48"}
    assert all(item["marker"] in {"DRAFT", "PLACEHOLDER"} for item in findings)


def test_reconciliation_does_not_start_forced_auto_calc_before_owned_excel(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    reconciled = tmp_path / "reconciled.xlsx"
    _lamd_fixture(source)

    finalizer.apply_lamd_truth_reconciliation(source, reconciled)

    calculation = load_workbook(reconciled, data_only=False).calculation
    assert calculation.calcMode is None
    assert calculation.forceFullCalc in (None, False)


def test_uncalculated_formula_cache_fails_closed(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    reconciled = tmp_path / "reconciled.xlsx"
    _lamd_fixture(source)
    receipt = finalizer.apply_lamd_truth_reconciliation(source, reconciled)

    with pytest.raises(finalizer.InvoiceFinalizationError, match="STALE_FORMULA_CACHE"):
        finalizer.verify_recalculated_invoice(
            reconciled,
            sheet_name="July 2026",
            invoice_number="2026-1004",
            expected_amount=100,
            expected_formula_sha256=receipt["pre_excel_formula_sha256"],
            balance_cell="G43",
        )


def test_formula_hash_detects_formula_mutation(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    mutated = tmp_path / "mutated.xlsx"
    _lamd_fixture(source)
    book = load_workbook(source, data_only=False)
    book["July 2026"]["G43"] = "=G40+G42+1"
    book.save(mutated)

    assert finalizer.formula_sha256(source) != finalizer.formula_sha256(mutated)


def test_verified_package_is_atomic_and_locator_compatible(tmp_path: Path) -> None:
    workbook = tmp_path / "verified.xlsx"
    _lamd_fixture(workbook)
    pdf = tmp_path / "verified.pdf"
    canvas = Canvas(str(pdf), pagesize=(612, 792))
    canvas.drawString(72, 720, "Verified invoice fixture")
    canvas.save()

    package = tmp_path / "root" / "live-arts-july"
    receipt = finalizer.publish_verified_package(
        package,
        workbook_path=workbook,
        pdf_path=pdf,
        specification={
            "client_ref": "live_arts_md",
            "invoice_number": "2026-1004",
            "service_period": "2026-07",
            "source_sheet": "July 2026",
            "amount": 100,
        },
        verification_receipt={
            "receipt_id": "invoice-w1-verify:test",
            "formula_freshness_receipt_id": "invoice-w1-formula:test",
            "status": "VERIFIED",
        },
    )

    assert receipt["status"] == "PUBLISHED_VERIFIED"
    assert (package / "invoice.xlsx").is_file()
    assert (package / "invoice.pdf").stat().st_size > 0
    located = invoice_artifact_locator.locate_invoice_artifacts(
        "live_arts_md",
        "2026-07",
        roots=[tmp_path / "root"],
    )
    assert located["status"] == "FOUND"
    assert located["canonical_candidate"]["invoice_number"] == "2026-1004"
    assert located["canonical_candidate"]["invoice_status"] == "finalized_verified"


def test_zero_byte_pdf_gets_distinct_failure_and_never_publishes(tmp_path: Path) -> None:
    workbook = tmp_path / "verified.xlsx"
    _lamd_fixture(workbook)
    pdf = tmp_path / "empty.pdf"
    pdf.touch()
    package = tmp_path / "package"

    with pytest.raises(finalizer.InvoiceFinalizationError, match="PDF_ZERO_BYTE"):
        finalizer.publish_verified_package(
            package,
            workbook_path=workbook,
            pdf_path=pdf,
            specification={
                "client_ref": "live_arts_md",
                "invoice_number": "2026-1004",
                "service_period": "2026-07",
                "source_sheet": "July 2026",
                "amount": 100,
            },
            verification_receipt={"status": "VERIFIED"},
        )

    assert not package.exists()


def test_pdf_semantic_marker_fails_before_atomic_publish(tmp_path: Path) -> None:
    workbook = tmp_path / "verified.xlsx"
    _lamd_fixture(workbook)
    pdf = tmp_path / "draft.pdf"
    canvas = Canvas(str(pdf), pagesize=(612, 792))
    canvas.drawString(72, 720, "DRAFT invoice - not final")
    canvas.save()
    package = tmp_path / "package"

    with pytest.raises(finalizer.InvoiceFinalizationError, match="PDF_SEMANTIC_MARKER_FOUND"):
        finalizer.publish_verified_package(
            package,
            workbook_path=workbook,
            pdf_path=pdf,
            specification={
                "client_ref": "live_arts_md",
                "invoice_number": "2026-1004",
                "service_period": "2026-07",
                "source_sheet": "July 2026",
                "amount": 100,
            },
            verification_receipt={"status": "VERIFIED"},
        )

    assert not package.exists()


def test_excel_com_workers_are_owned_bounded_and_never_kill_ambient_excel() -> None:
    root = Path(__file__).resolve().parents[1]
    recalc = (root / "scripts" / "recalculate_invoice_with_excel.ps1").read_text(
        encoding="utf-8"
    )
    export = (root / "scripts" / "export_invoice_pdf_with_excel.ps1").read_text(
        encoding="utf-8"
    )
    combined = recalc + export

    assert "CalculateFullRebuild" in recalc
    assert "CalculationState" in recalc
    assert "AutomationSecurity = 3" in combined
    assert "DisplayAlerts = $false" in combined
    assert "AskToUpdateLinks = $false" in combined
    assert "LinkSources" in recalc
    assert "ExportAsFixedFormat" in export
    assert "[string]$PrintArea" in export
    assert "$worksheet.PageSetup.PrintArea = $PrintArea" in export
    assert "Stop-Process" not in combined
    assert "taskkill" not in combined.lower()
    assert "Get-Process" not in combined


def test_pdf_export_applies_bounded_print_area_without_mutating_workbook(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    workbook = tmp_path / "invoice.xlsx"
    output = tmp_path / "candidate-b.pdf"
    _lamd_fixture(workbook)
    source_sha256 = _sha256(workbook)
    captured: dict[str, object] = {}

    def fake_excel_worker(script: Path, args: list[str], *, timeout: int) -> dict[str, str]:
        captured.update({"script": script, "args": args, "timeout": timeout})
        temporary = Path(args[args.index("-OutputPath") + 1])
        temporary.write_bytes(b"%PDF-1.4\n%%EOF\n")
        return {"excel_version": "16.0", "print_area": "$A$1:$H$48"}

    monkeypatch.setattr(finalizer, "_windows_path", lambda path: str(path))
    monkeypatch.setattr(finalizer, "_run_excel_worker", fake_excel_worker)
    monkeypatch.setattr(
        finalizer,
        "_validate_pdf",
        lambda path: {
            "sha256": _sha256(path),
            "size": path.stat().st_size,
            "page_count": 1,
        },
    )

    receipt = finalizer.export_invoice_pdf_with_excel(
        workbook,
        sheet_name="July 2026",
        output_path=output,
        print_area="$A$1:$H$48",
    )

    assert captured["args"][-2:] == ["-PrintArea", "$A$1:$H$48"]
    assert receipt["print_area_applied"] == "$A$1:$H$48"
    assert receipt["workbook_sha256_before"] == source_sha256
    assert receipt["workbook_sha256_after"] == source_sha256
    assert receipt["workbook_unchanged"] is True
    assert _sha256(workbook) == source_sha256
    assert output.is_file()


@pytest.mark.parametrize(
    "print_area",
    ["A1:H48", "$A$0:$H$48", "$A$1:$H$48, $A$50:$H$56", "July 2026!$A$1:$H$48"],
)
def test_pdf_export_rejects_unbounded_print_area_before_owner_process(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, print_area: str
) -> None:
    workbook = tmp_path / "invoice.xlsx"
    _lamd_fixture(workbook)
    monkeypatch.setattr(
        finalizer,
        "_run_excel_worker",
        lambda *_args, **_kwargs: pytest.fail("owner process must not run"),
    )

    with pytest.raises(finalizer.InvoiceFinalizationError, match="PRINT_AREA_INVALID"):
        finalizer.export_invoice_pdf_with_excel(
            workbook,
            sheet_name="July 2026",
            output_path=tmp_path / "preview.pdf",
            print_area=print_area,
        )


def test_excel_adapter_rejects_source_output_alias_before_process_call(tmp_path: Path) -> None:
    workbook = tmp_path / "invoice.xlsx"
    _lamd_fixture(workbook)

    with pytest.raises(finalizer.InvoiceFinalizationError, match="RECALC_OUTPUT_MUST_DIFFER"):
        finalizer.recalculate_with_excel(workbook, workbook)
