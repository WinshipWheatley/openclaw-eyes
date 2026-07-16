"""Read-only St. Anne's workbook-to-worklog truth-drift detector."""

from __future__ import annotations

import argparse
import hashlib
import json
from collections.abc import Mapping
from pathlib import Path
from typing import Any

import openpyxl


SCHEMA_VERSION = "st_annes_invoice_truth_drift_v0"
DEFAULT_MANIFEST_PATH = Path(
    "/mnt/e/openclaw/codex_mac_bridge/from-codex-mac/invoice_handoffs/"
    "st-annes-2026-06-real-pdf-20260704T231607/invoice_manifest.json"
)
DEFAULT_HYGIENE_PATH = Path("generated/read_models/st_annes_work_log_hygiene.json")
DEFAULT_OUTPUT_PATH = Path("generated/read_models/st_annes_invoice_truth_drift.json")


def _json_object(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalized(value: object) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _as_number(value: object) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace("$", "").replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _extract_workbook_truth(workbook_path: Path, sheet_name: str) -> tuple[dict[str, Any], list[str]]:
    blockers: list[str] = []
    try:
        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - source failures become bounded blockers.
        return {}, [f"workbook_parse_failed:{type(exc).__name__}"]
    try:
        if sheet_name not in workbook.sheetnames:
            return {}, [f"declared_sheet_missing:{sheet_name}"]
        sheet = workbook[sheet_name]
        rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    header_index = -1
    for index, row in enumerate(rows):
        second = _normalized(row[1] if len(row) > 1 else None)
        third = _normalized(row[2] if len(row) > 2 else None)
        if second == "item #" and third == "description":
            header_index = index
            break
    if header_index < 0:
        blockers.append("service_table_header_missing")

    services: list[dict[str, Any]] = []
    if header_index >= 0:
        for row in rows[header_index + 1 :]:
            item_number = _as_number(row[1] if len(row) > 1 else None)
            description = str(row[2] or "").strip() if len(row) > 2 else ""
            quantity = _as_number(row[3] if len(row) > 3 else None)
            unit_price = _as_number(row[4] if len(row) > 4 else None)
            if item_number is None:
                if services:
                    break
                continue
            if not description or quantity is None or unit_price is None:
                blockers.append(f"service_row_incomplete:{int(item_number)}")
                continue
            services.append(
                {
                    "item_number": int(item_number),
                    "description": description,
                    "quantity": quantity,
                    "unit_price": unit_price,
                    "line_amount": round(quantity * unit_price, 2),
                }
            )

    total_due: float | None = None
    for row in rows:
        for index, value in enumerate(row):
            if _normalized(value) != "total due":
                continue
            for candidate in row[index + 1 :]:
                total_due = _as_number(candidate)
                if total_due is not None:
                    break
            break
        if total_due is not None:
            break
    if total_due is None:
        blockers.append("total_due_missing")

    return {
        "source_sheet": sheet_name,
        "service_count": len(services),
        "services": services,
        "total_due": total_due,
    }, blockers


def _base_receipt(
    manifest_path: Path,
    hygiene_path: Path,
    *,
    generated_at: str,
) -> dict[str, Any]:
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": generated_at,
        "client_ref": "st_annes",
        "service_period": "2026-06",
        "status": "SOURCE_UNAVAILABLE",
        "manifest_path": manifest_path.as_posix(),
        "hygiene_path": hygiene_path.as_posix(),
        "workbook_truth": {},
        "mirror_truth": {},
        "missing_items": [],
        "blockers": [],
        "authority_boundary": {
            "email_send_allowed": False,
            "ledger_posting_allowed": False,
            "workbook_mutation_allowed": False,
            "pdf_export_allowed": False,
            "paid_marking_allowed": False,
        },
        "machine_proof": {
            "manifest_read_performed": False,
            "workbook_read_performed": False,
            "workbook_hash_verified": False,
            "workbook_mutation_performed": False,
            "ledger_mutation_performed": False,
            "email_send_performed": False,
            "pdf_export_performed": False,
            "paid_marking_performed": False,
        },
    }


def build_truth_drift(
    manifest_path: Path,
    hygiene_path: Path,
    *,
    generated_at: str,
) -> dict[str, Any]:
    manifest_path = Path(manifest_path)
    hygiene_path = Path(hygiene_path)
    receipt = _base_receipt(manifest_path, hygiene_path, generated_at=generated_at)
    manifest = _json_object(manifest_path)
    hygiene = _json_object(hygiene_path)
    blockers: list[str] = []
    if not manifest:
        blockers.append("manifest_missing_or_invalid")
    else:
        receipt["machine_proof"]["manifest_read_performed"] = True

    raw_confirmed_ids = hygiene.get("business_confirmed_ready_event_ids")
    if not isinstance(raw_confirmed_ids, list):
        blockers.append("mirror_confirmed_ids_invalid")
        confirmed_ids: list[str] = []
    else:
        confirmed_ids = sorted(
            {str(item).strip() for item in raw_confirmed_ids if str(item).strip()}
        )
    receipt["mirror_truth"] = {
        "confirmed_event_count": len(confirmed_ids),
        "confirmed_event_ids": confirmed_ids,
    }

    if not manifest:
        receipt["blockers"] = blockers
        return receipt

    workbook_path = manifest_path.parent / "invoice.xlsx"
    receipt["workbook_path"] = workbook_path.as_posix()
    if not workbook_path.is_file():
        blockers.append("workbook_missing")
        receipt["blockers"] = blockers
        return receipt

    actual_hash = _sha256(workbook_path)
    expected_hash = str(manifest.get("package_workbook_sha256") or "").strip().lower()
    receipt["workbook_sha256"] = actual_hash
    if not expected_hash or actual_hash != expected_hash:
        blockers.append("workbook_hash_mismatch")
        receipt["status"] = "HASH_MISMATCH"
        receipt["blockers"] = blockers
        return receipt
    receipt["machine_proof"]["workbook_hash_verified"] = True

    sheet_name = str(manifest.get("source_sheet") or "").strip()
    if not sheet_name:
        blockers.append("manifest_source_sheet_missing")
        receipt["blockers"] = blockers
        return receipt
    workbook_truth, workbook_blockers = _extract_workbook_truth(workbook_path, sheet_name)
    blockers.extend(workbook_blockers)
    if workbook_truth:
        receipt["machine_proof"]["workbook_read_performed"] = True
        workbook_truth["invoice_number"] = str(manifest.get("invoice_number") or "")
        workbook_truth["manifest_amount"] = _as_number(manifest.get("amount"))
        workbook_truth["invoice_status"] = str(manifest.get("status") or "")
        workbook_truth["send_receipt_present"] = bool(manifest.get("latest_send_receipt_path"))
        receipt["workbook_truth"] = workbook_truth
        manifest_amount = workbook_truth["manifest_amount"]
        total_due = workbook_truth["total_due"]
        if manifest_amount is not None and total_due is not None and round(manifest_amount, 2) != round(total_due, 2):
            blockers.append("manifest_total_mismatch")

    if blockers:
        receipt["blockers"] = blockers
        return receipt

    service_count = int(receipt["workbook_truth"]["service_count"])
    confirmed_count = int(receipt["mirror_truth"]["confirmed_event_count"])
    if service_count != confirmed_count:
        receipt["status"] = "DRIFT_DETECTED"
        receipt["missing_items"] = [
            "Reconcile workbook billables into the work-log mirror"
        ]
    else:
        receipt["status"] = "IN_SYNC"
    receipt["blockers"] = []
    return receipt


def write_truth_drift(payload: Mapping[str, Any], output_path: Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(dict(payload), indent=2, sort_keys=True, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    return output_path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST_PATH)
    parser.add_argument("--hygiene", type=Path, default=DEFAULT_HYGIENE_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    parser.add_argument("--generated-at", required=True)
    args = parser.parse_args(argv)
    payload = build_truth_drift(
        args.manifest,
        args.hygiene,
        generated_at=args.generated_at,
    )
    write_truth_drift(payload, args.output)
    print(json.dumps(payload, indent=2, sort_keys=True, ensure_ascii=False))
    return 0 if payload["status"] in {"DRIFT_DETECTED", "IN_SYNC"} else 2


if __name__ == "__main__":
    raise SystemExit(main())
